import csv, os
from pathlib import Path
import openpyxl

def catalog_file(filepath):
    records = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        fname = Path(filepath).name
        year = next((int(y) for y in ["2015","2016","2017","2018","2019","2020","2021","2022","2023"] if y in fname), None)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=6, values_only=True))
            n_rows = sum(1 for _ in ws.iter_rows())
            col_names = [str(c) if c is not None else "" for c in (rows[0] if rows else [])]
            records.append({
                "filename": fname,
                "year": year,
                "sheet": sheet_name,
                "n_rows": n_rows,
                "n_cols": ws.max_column,
                "has_location_col": "Location" in col_names,
                "has_year_col": "Year" in col_names,
                "columns_preview": " | ".join(col_names[:8]),
            })
        wb.close()
    except Exception as e:
        records.append({"filename": Path(filepath).name, "year": None, "sheet": "ERROR", "n_rows": None, "n_cols": None, "has_location_col": None, "has_year_col": None, "columns_preview": str(e)})
    return records

data_dir = Path("data/raw/years")
files = sorted(data_dir.glob("**/*.xlsx")) + sorted(data_dir.glob("**/*.xls"))
print(f"Found {len(files)} files")

all_records = []
for f in files:
    print(f"  {f.name}")
    all_records.extend(catalog_file(f))

Path("docs").mkdir(exist_ok=True)
with open("docs/file_inventory.csv", "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["filename","year","sheet","n_rows","n_cols","has_location_col","has_year_col","columns_preview"])
    writer.writeheader()
    writer.writerows(all_records)

print(f"\nDone. {len(all_records)} sheets cataloged → docs/file_inventory.csv")
